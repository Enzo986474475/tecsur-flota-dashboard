# 1_Operacion.py
from datetime import date, timedelta
from pathlib import Path
import re
import unicodedata as ud
import pandas as pd
import streamlit as st
import altair as alt
from pandas.io.formats.style import Styler


from components.ui import header
from lib.config import load_settings
from lib.data import load_excel
from lib.sync import sync_from_settings
from urllib.parse import quote


# ====== Sincroniza últimos datos (OneDrive) ======
settings = load_settings()
sync_from_settings(settings)

# Ruta de Control de Flota con fallback
FLOTA_PATH = (
    settings.get("FLOTA_PATH")
    or settings.get("FLOTA_XLSX")
    or settings.get("FLOTA_FILE")
    or r"C:\Users\Enzo Morote\TECSUR S.A\Gestión Flota EV - Enzo - Documentos\Control-de-Flota-Vehicular-Tecsur 12.xlsx"
)




st.set_page_config(page_title="Tecsur · Flota EV — Operación", page_icon="📊", layout="wide")

# ============== Encabezado ==============
#settings = load_settings()
header("Operación")

# ============== Sidebar ==============
with st.sidebar:
    st.subheader("Filtros")
    fecha_ini, fecha_fin = st.date_input(
        "Rango de fechas",
        value=(date.today().replace(day=1), date.today())
    )
    placa_filtro = st.text_input("Placa / Vehículo (opcional)")
      
    st.divider()
    st.subheader("Acciones")

    if st.button("🔄 Actualizar datos", use_container_width=True):
        # Re-sincroniza (OneDrive/SharePoint) si aplica
        sync_from_settings(settings)
        # Limpia el caché de dataframes
        st.cache_data.clear()
        # Vuelve a ejecutar la página
        st.rerun()


# ============== Recorridos ==============
rec_path = settings.get("data", {}).get("recorridos_maestro", "")
df = load_excel(rec_path)

if df.empty:
    st.info("No pude leer el archivo de recorridos. Verifica la ruta en `settings.yaml`.")
    st.stop()

df.columns = [str(c).strip() for c in df.columns]

def canon(s: str) -> str:
    s = str(s).strip().lower()
    s = ''.join(ch for ch in ud.normalize('NFD', s) if ud.category(ch) != 'Mn')
    s = s.replace('(', ' ').replace(')', ' ').replace('/', ' ').replace('-', ' ')
    s = re.sub(r'[^a-z0-9\s]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def find_col(keys: list[str]) -> str | None:
    cand = [(c, canon(c)) for c in df.columns]
    for c, cc in cand:
        for k in keys:
            if k in cc:
                return c
    return None

col_fecha   = find_col(["fecha", "date"])
col_km      = find_col(["km", "kilometraje", "distancia km", "distancia"])
col_placa   = find_col(["placa", "vehiculo", "unidad"])
col_familia = find_col(["familia", "tipo", "segmento", "categoria"])

if not (col_fecha and col_km):
    st.error("No encontré columnas de **Fecha** y/o **Km** en el Excel de recorridos.")
    st.write("Columnas disponibles:", list(df.columns))
    st.stop()

tmp = df.copy()
tmp[col_fecha] = pd.to_datetime(tmp[col_fecha], errors="coerce").dt.date
tmp = tmp[(tmp[col_fecha] >= fecha_ini) & (tmp[col_fecha] <= fecha_fin)]
if placa_filtro and col_placa:
    tmp = tmp[tmp[col_placa].astype(str).str.contains(placa_filtro.strip(), case=False, na=False)]

km_total = float(tmp[col_km].fillna(0).sum())
unid_sin_mov = int((tmp.groupby(col_placa)[col_km].sum() == 0).sum()) if col_placa else 0



st.divider()

# =====================================================================
# ===================  Helpers y lecturas para Check list  =============
# =====================================================================

INSPECCIONES_PATH = Path(r"C:\Users\Enzo Morote\TECSUR S.A\Gestión Flota EV - Enzo - Documentos\1.-Check list\Inspecciones.xlsx")
FLOTA_PATH        = Path(FLOTA_PATH) if not isinstance(FLOTA_PATH, Path) else FLOTA_PATH   # usa el ya definido arriba
FLOTA_OPER_PATH   = Path(r"C:\Users\Enzo Morote\TECSUR S.A\Gestión Flota EV - Enzo - Documentos\Flota_Operciones.xlsx")




# Carpeta (vista) — siempre abre la carpeta
BASE_INSP_FOLDER_VIEW = (
    "https://tecsurpe.sharepoint.com/sites/GestinFlotaEV-Enzo/"
    "Documentos%20compartidos/Forms/AllItems.aspx?"
    "id=%2Fsites%2FGestinFlotaEV%2DEnzo%2FDocumentos%20compartidos%2F1%2E%2DCheck%20list%2FInspecciones"
)


SP_HOST = "https://tecsurpe.sharepoint.com"

def sp_normalize(url: str) -> str:
    """Convierte /sites/... o sites/... en absoluta y escapa espacios.
       Si ya es http(s):// la deja tal cual (solo escapa espacios)."""
    if not url:
        return ""
    u = str(url).strip()

    # Ya absoluta
    if u.startswith("http://") or u.startswith("https://"):
        return u.replace(" ", "%20")

    # Acepta 'sites/...' o '/sites/...'
    if u.startswith("sites/"):
        u = "/" + u
    if u.startswith("/sites/"):
        u = SP_HOST + u

    return u.replace(" ", "%20")

def sp_web_view(u: str) -> str:
    return "" if not u else (u if "?" in u else u + "?web=1")




# Base directa a archivo (intentaremos ambas por si la carpeta se llama con o sin la 's')
BASE_FILE_INSPECCIONES = "https://tecsurpe.sharepoint.com/sites/GestinFlotaEV-Enzo/Documentos%20compartidos/1.-Check%20list/Inspecciones/"
BASE_FILE_INPECCIONES  = "https://tecsurpe.sharepoint.com/sites/GestinFlotaEV-Enzo/Documentos%20compartidos/1.-Check%20list/Inpecciones/"

INSPECCIONES_DIR = Path(r"C:\Users\Enzo Morote\TECSUR S.A\Gestión Flota EV - Enzo - Documentos\1.-Check list\Inspecciones")


def _local_pdf_candidates(placa: str) -> list[Path]:
    return [
        INSPECCIONES_DIR / f"Check-list-Unidades_{placa}.pdf",
        INSPECCIONES_DIR / f"Check list Unidades_{placa}.pdf",
    ]

def find_local_pdf(placa: str) -> Path | None:
    for p in _local_pdf_candidates(placa):
        if p.exists():
            return p
    return None
##Continuamos con mas helpers--------------

SP_HOST = "https://tecsurpe.sharepoint.com"

def sp_normalize(url: str) -> str:
    """Devuelve URL absoluta a SharePoint y con espacios escapados."""
    if not url:
        return ""
    url = str(url).strip()

    # Acepta  sites/...  y  /sites/... 
    if url.startswith("sites/"):
        url = "/" + url
    if url.startswith("/sites"):
        url = SP_HOST + url

    # Escapa espacios si vinieran sin encode
    if " " in url:
        url = url.replace(" ", "%20")
    return url

def sp_web_view(url: str) -> str:
    """Fuerza el modo visor web si no hay query."""
    if not url:
        return ""
    return url if "?" in url else url + "?web=1"



_pat_placa = re.compile(r"\b([A-Z]{3})[- ]?(\d{3})\b", re.I)
def norm_placa(s: str) -> str:
    s = str(s or "").upper().strip().replace("—", "-").replace("–", "-")
    m = _pat_placa.search(s)
    return f"{m.group(1)}-{m.group(2)}" if m else s

def excel_col_to_idx(col_letters: str) -> int:
    """A -> 0, Z -> 25, AA -> 26, AR -> 43, ..."""
    col_letters = col_letters.strip().upper()
    n = 0
    for ch in col_letters:
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n - 1

@st.cache_data(show_spinner=False)
def read_detalle(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Detalle", header=2)  # header fila 3
    df.columns = [str(c).strip() for c in df.columns]
    # columna de fecha (si existe)
    fcol = next((c for c in df.columns if canon(c) in {canon("Datos Generales_Fecha"), "fecha"}), None)
    if fcol:
        df[fcol] = pd.to_datetime(df[fcol], errors="coerce").dt.date
    return df

@st.cache_data(show_spinner=False)
def read_flota(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Flota", header=7)  # fila 8
    df = df.iloc[:, 2:].copy()  # desde C
    df.columns = [str(c).strip() for c in df.columns]
    return df



@st.cache_data(show_spinner=False)
def read_flota_oper_info(path: Path) -> tuple[dict, dict]:
    """Devuelve (fechas_map, links_map) desde 'Flota inspeccionada'.
    Lee hiperlinks reales del Excel (no el texto mostrado)."""
    if not path.exists():
        return {}, {}

    import openpyxl

    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return {}, {}
    if "Flota inspeccionada" not in wb.sheetnames:
        return {}, {}
    ws = wb["Flota inspeccionada"]

    # Lee encabezados
    headers = {}
    for j, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1):
        headers[j] = str(cell.value).strip() if cell.value is not None else ""

    # Ubica columnas
    def hkey(s): 
        return re.sub(r"\s+", " ", str(s).strip().lower())
    col_idx_placa = next((j for j, h in headers.items() if hkey(h) in {"placa vehiculo", "placa vehículo", "placa"}), None)
    col_idx_fecha = next((j for j, h in headers.items() if hkey(h) in {"fecha inspección", "fecha inspeccion", "fecha"}), None)
    col_idx_link  = next((j for j, h in headers.items() if hkey(h) in {"link", "ver inspección", "ver inspeccion"}), None)

    if not col_idx_placa:
        return {}, {}

    fechas_map, links_map = {}, {}

    for row in ws.iter_rows(min_row=2):  # datos
        placa_raw = row[col_idx_placa-1].value
        if not placa_raw:
            continue
        p = norm_placa(placa_raw)

        # Fecha formateada (si existe)
        if col_idx_fecha:
            fv = row[col_idx_fecha-1].value
            try:
                dt = pd.to_datetime(fv, errors="coerce")
                if pd.notna(dt):
                    fechas_map[p] = dt.strftime("%d/%m/%Y")
            except Exception:
                pass

        # URL real del hyperlink (si existe)
        url = ""
        if col_idx_link:
            cell = row[col_idx_link-1]
            # 1) Si hay hyperlink, úsalo
            if getattr(cell, "hyperlink", None) and getattr(cell.hyperlink, "target", ""):
                url = cell.hyperlink.target
            else:
                # 2) A veces el valor guarda la fórmula HYPERLINK("...","...")
                val = str(cell.value or "")
                m = re.search(r'HYPERLINK\("([^"]+)"\s*,', val, flags=re.I)
                if m:
                    url = m.group(1)

        if url:
            links_map[p] = url

    wb.close()
    return fechas_map, links_map


def find_col_in(df_any: pd.DataFrame, keys: list[str]) -> str | None:
    """Busca una columna cuyo nombre (normalizado) contenga alguno de los keys."""
    if df_any is None or df_any.empty:
        return None
    def _canon(s: str) -> str:
        s = str(s or "").strip().lower()
        s = ''.join(ch for ch in ud.normalize('NFD', s) if ud.category(ch) != 'Mn')
        s = s.replace('(', ' ').replace(')', ' ').replace('/', ' ').replace('-', ' ')
        s = re.sub(r'[^a-z0-9\s]+', ' ', s)
        s = re.sub(r'\s+', ' ', s).strip()
        return s
    keys_c = [_canon(k) for k in keys]
    for c in df_any.columns:
        cc = _canon(c)
        if any(k in cc for k in keys_c):
            return c
    return None



def find_col_in(df_any: pd.DataFrame, keys: list[str]) -> str | None:
    """Busca una columna cuyo nombre (normalizado) contenga alguno de los keys."""
    if df_any is None or df_any.empty:
        return None
    def _canon(s: str) -> str:
        s = str(s or "").strip().lower()
        s = ''.join(ch for ch in ud.normalize('NFD', s) if ud.category(ch) != 'Mn')
        s = s.replace('(', ' ').replace(')', ' ').replace('/', ' ').replace('-', ' ')
        s = re.sub(r'[^a-z0-9\s]+', ' ', s)
        s = re.sub(r'\s+', ' ', s).strip()
        return s
    keys_c = [_canon(k) for k in keys]
    for c in df_any.columns:
        cc = _canon(c)
        if any(k in cc for k in keys_c):
            return c
    return None


def _guess_pdf_link(placa: str) -> str:
    p = norm_placa(placa)
    # probar ambos nombres y ambas carpetas
    candidates = []
    for base in (BASE_FILE_INSPECCIONES, BASE_FILE_INPECCIONES):
        candidates.append(base + quote(f"Check-list-Unidades_{p}.pdf"))
        candidates.append(base + quote(f"Check list Unidades_{p}.pdf"))
    # devolvemos la primera candidata en modo web
    return sp_web_view(sp_normalize(candidates[0]))




def _estado_general(row: pd.Series, cat_cols: list[str]) -> str:
    n = sum(1 for c in cat_cols if row.get(c) == "❌")
    if n == 0:
        return "🟢 Conforme"
    elif n <= 2:
        return "🟡 Regular"
    else:
        return "🔴 Crítico"


def build_checklist_table(df_det: pd.DataFrame, df_flota: pd.DataFrame,
                          fechas_map: dict, fecha_ini: date, fecha_fin: date,
                          placa_filter: str | None = None) -> Styler | pd.DataFrame:
    # columnas clave
    placa_col = next((c for c in df_det.columns if canon(c) in {"placa", "vehiculo", "vehículo", "unidad"}), None)
    fecha_col = next((c for c in df_det.columns if canon(c) in {canon("Datos Generales_Fecha"), "fecha"}), None)
    if not placa_col:
        st.error("No se encontró la columna **Placa** en Inspecciones → Detalle."); st.stop()

    df = df_det.copy()
    if fecha_col:
        df = df[df[fecha_col].between(fecha_ini, fecha_fin, inclusive="both")]
    if placa_filter:
        df = df[df[placa_col].astype(str).str.contains(str(placa_filter).strip(), case=False, na=False)]
    df["_placa"] = df[placa_col].map(norm_placa)

    # rangos por posición (C–AR)
    headers = list(df.columns)
    def names_in_range(col_a: str, col_b: str) -> list[str]:
        a = excel_col_to_idx(col_a); b = excel_col_to_idx(col_b)
        a = max(0, a); b = min(len(headers)-1, b)
        if a > b: return []
        return [headers[i] for i in range(a, b+1)]

    grupos = {
        "Documentos":  names_in_range("C", "G"),
        "Implementos": names_in_range("H", "R"),
        "Luces":       names_in_range("S", "V"),
        "Neumáticos":  names_in_range("W", "X"),
        "Unidad":      names_in_range("Y", "AR"),
    }

    # cliente asignado
    placa_col_flota = next((c for c in df_flota.columns if canon(c).startswith("placa")), None)
    cliente_col     = next((c for c in df_flota.columns if canon(c) == "cliente asignado"), None)
    map_cliente = {}
    if placa_col_flota and cliente_col:
        dff = df_flota[[placa_col_flota, cliente_col]].dropna(subset=[placa_col_flota]).copy()
        dff["_placa"] = dff[placa_col_flota].map(norm_placa)
        map_cliente = dff.drop_duplicates("_placa").set_index("_placa")[cliente_col].to_dict()

    # última fila por placa
    ult = df.sort_values(fecha_col if fecha_col else df.index).drop_duplicates("_placa", keep="last")

    # construir tabla
    filas = []
    tips_rows = []
    detalles_fallback = []
    for _, row in ult.iterrows():
        p = row["_placa"]
        registro = {"Placa": p, "Cliente Asignado": map_cliente.get(p, "")}
        tips = {"Placa": "", "Cliente Asignado": ""}
        detalle_txt = []

        for cat, cols_cat in grupos.items():
            vals = [str(row.get(c, "")).strip().lower() for c in cols_cat]
            hay_mal = any(v == "mal" for v in vals)
            hay_ok  = any(v in ("ok", "conforme") for v in vals)

            if hay_mal:
                registro[cat] = "❌"
                malos = [c for c in cols_cat if str(row.get(c, "")).strip().lower() == "mal"]
                fecha_txt = fechas_map.get(p, "")
                det = "; ".join(f"{c}: Mal" for c in malos)
                tips[cat] = f"{det}{(' · Checklist ' + fecha_txt) if fecha_txt else ''}"
                detalle_txt.append(tips[cat])
            elif hay_ok:
                registro[cat] = "✅"
                tips[cat] = ""
            else:
                registro[cat] = ""  # vacío
                tips[cat] = ""

        filas.append(registro)
        tips_rows.append(tips)
        detalles_fallback.append("; ".join([d for d in detalle_txt if d]))

    out = pd.DataFrame(filas).sort_values("Placa").reset_index(drop=True)
    tips_df = pd.DataFrame(tips_rows).loc[out.index, out.columns]

    # Estilo tablero
    header_bg = "#1f4e78"; header_fg = "#ffffff"
    body_bg   = "#fff3b0"; border    = "1px solid #dddddd"

    styler = out.style.set_properties(**{
        "text-align": "center", "background-color": body_bg, "border": border
    }).set_table_styles([
        {"selector": "thead th", "props": [("background-color", header_bg), ("color", header_fg),
                                           ("text-align", "center"), ("border", border)]},
        {"selector": "tbody td", "props": [("text-align", "center"), ("border", border)]},
    ]).format(lambda v: v)

    # ocultar índice (compatibilidad)
    try:
        styler = styler.hide_index()
    except Exception:
        try: styler = styler.hide(axis="index")
        except Exception: pass

    # tooltips por celda solo si la versión lo soporta
    if hasattr(styler, "set_tooltips"):
        styler = styler.set_tooltips(tips_df)
        return styler
    else:
        out_fallback = out.copy()
        out_fallback["Detalle"] = [d if d else "" for d in detalles_fallback]
        if not any(bool(x) for x in out_fallback["Detalle"]):
            out_fallback = out_fallback.drop(columns=["Detalle"])
        return out_fallback

# ============== Tabs ==============
tab_resumen, tab_detalle, tab_check, tab_export = st.tabs(["Resumen", "Detalle", "Lista de verificación", "Exportar"])



with tab_resumen:
    st.subheader("Resumen")

    # ===== KPIs (seguros) =====
   
    # ===== Flota Tecsur (Parque — Placas por Familia × Combustible) =====
    st.markdown("#### Flota Tecsur")

    try:
        df_flota_parque = read_flota(FLOTA_PATH)  # hoja "Flota" (header=7)
    except Exception as e:
        df_flota_parque = pd.DataFrame()
        st.warning(f"No pude leer 'Flota' para el cuadro de parque: {e}")

    if df_flota_parque.empty:
        st.info("No encontré datos en **Flota**.")
    else:
        # Columnas clave
        col_fam   = find_col_in(df_flota_parque, ["familia", "segmento", "tipo"])
        col_comb  = find_col_in(df_flota_parque, ["combustible", "motorizacion", "motorización", "fuel"])
        col_placa = find_col_in(df_flota_parque, ["placa"])

        if not col_fam or not col_comb:
            st.info("No encontré columnas de **Familia** y/o **Combustible** en la hoja *Flota*.")
        else:
            dfp = df_flota_parque.copy()
            dfp[col_fam]  = dfp[col_fam].astype(str).str.strip()
            dfp[col_comb] = dfp[col_comb].astype(str).str.strip().str.upper()

            def _norm_comb(x: str) -> str:
                x0 = x.replace("ELECTRICO", "ELÉCTRICO").replace("HIBRIDA", "HÍBRIDA")
                x0 = x0.replace("HÍBRIDO", "HÍBRIDA")
                # Normalizamos GASOHOL/GASOLINA a GASOLINA (tu header anterior mostraba GASOLINA)
                x0 = x0.replace("GASOHOL", "GASOLINA")
                return x0

            dfp[col_comb] = dfp[col_comb].map(_norm_comb)

            # Evitar duplicados por placa
            if col_placa:
                dfp["_PLACA_NORM"] = dfp[col_placa].astype(str).str.upper().str.strip()
                dfp = dfp.drop_duplicates(subset=["_PLACA_NORM"])

            piv = (
                dfp[[col_fam, col_comb]]
                .dropna(subset=[col_fam, col_comb])
                .assign(_n=1)
                .pivot_table(index=col_fam, columns=col_comb, values="_n",
                             aggfunc="sum", fill_value=0)
            )

            # Orden de columnas como te gustó
            ordered = [c for c in ["DIÉSEL", "DIESEL", "ELÉCTRICO", "GASOLINA", "HÍBRIDA"] if c in piv.columns]
            # Unificamos DIESEL → DIÉSEL si existe
            if "DIESEL" in piv.columns and "DIÉSEL" not in piv.columns:
                piv = piv.rename(columns={"DIESEL": "DIÉSEL"})
                ordered = [c.replace("DIESEL", "DIÉSEL") for c in ordered]
            other   = [c for c in piv.columns if c not in ordered]
            piv = piv[ordered + other] if ordered else piv

            # Totales y orden por total
            piv["Familia total"] = piv.sum(axis=1)
            piv = piv.sort_values("Familia total", ascending=False)

            # Quitar nombre del eje columnas (evita cabecera extra)
            piv.columns.name = None

            # ===== KPIs de Flota (se muestran antes del cuadro) =====
            # Total de placas (preferimos conteo por placa; si no hubiera, usamos el total del pivote)
            if col_placa:
                flota_total = int(dfp[col_placa].astype(str).str.upper().str.strip().nunique())
            else:
                flota_total = int(piv["Familia total"].sum())

            # Sumas por tipo de motorización desde el pivote
            flota_ev       = int(piv["ELÉCTRICO"].sum()) if "ELÉCTRICO" in piv.columns else 0
            flota_hibrida  = int(piv["HÍBRIDA"].sum())   if "HÍBRIDA"   in piv.columns else 0

            # La columna diésel puede venir como DIESEL o DIÉSEL; tomamos la que exista
            diesel_col = None
            if "DIÉSEL" in piv.columns: diesel_col = "DIÉSEL"
            elif "DIESEL" in piv.columns: diesel_col = "DIESEL"

            suma_diesel   = int(piv[diesel_col].sum()) if diesel_col else 0
            suma_gasolina = int(piv["GASOLINA"].sum()) if "GASOLINA" in piv.columns else 0
            flota_ci = suma_diesel + suma_gasolina

          
            # ===== KPI CARDS — reemplaza el bloque m1..m4.metric por este =====
            # Requiere que ya existan: flota_total, flota_ev, flota_hibrida, flota_ci

            kpi_css = """
            <style>
            .kpi-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px; }
            @media (max-width: 1200px) { .kpi-grid { grid-template-columns: repeat(2, 1fr); } }
            .kpi {
                position: relative; background: #FFFFFF; border: 1px solid #eef2f7;
                border-radius: 18px; padding: 18px 18px 14px 18px;
                box-shadow: 0 6px 18px rgba(17,24,39,.08);
            }
            .kpi::before{
                content:""; position:absolute; left:0; right:0; top:0; height:4px;
                background: var(--accent, #2D5252);
                border-top-left-radius: 18px; border-top-right-radius: 18px;
            }
            .kpi-top { display:flex; align-items:center; gap:10px; margin-bottom:6px; }
            .kpi-icon {
                width:30px; height:30px; border-radius:999px;
                display:flex; align-items:center; justify-content:center;
                font-size:18px; background: var(--accent-bg, #E6F4F1); color: var(--accent,#2D5252);
            }
            .kpi-label { font-size: 13px; color: #475569; font-weight: 600; }
            .kpi-value { font-size: 48px; line-height: 1.0; font-weight: 800; color:#0f172a; letter-spacing: -0.5px; }
            </style>
            """

            def _fmt(n):
                try:
                    return f"{int(n):,}".replace(",", ",")
                except:
                    return str(n)

            def kpi_card(icon: str, label: str, value, accent: str, accent_bg: str):
                return f"""
                <div class="kpi" style="--accent:{accent};--accent-bg:{accent_bg};">
                <div class="kpi-top">
                    <div class="kpi-icon">{icon}</div>
                    <div class="kpi-label">{label}</div>
                </div>
                <div class="kpi-value">{_fmt(value)}</div>
                </div>
                """

            # Paleta (acento y fondo suave del icono)
            ACC_TOTAL = "#2D5252"; BG_TOTAL = "#E6F4F1"
            ACC_EV    = "#0EA5A0"; BG_EV    = "#ECFDF5"
            ACC_HIB   = "#22C55E"; BG_HIB   = "#F0FDF4"
            ACC_CI    = "#3B82F6"; BG_CI    = "#EEF2FF"

            # Render: fila de 4 tarjetas
            cards_html = """
            <div class="kpi-grid">
            {card1}
            {card2}
            {card3}
            {card4}
            </div>
            """.format(
            card1=kpi_card("🚘", "Flota Total", flota_total, ACC_TOTAL, BG_TOTAL),
            card2=kpi_card("⚡", "Flota Eléctrica (EV)", flota_ev, ACC_EV, BG_EV),
            card3=kpi_card("🌿", "Flota Híbrida", flota_hibrida, ACC_HIB, BG_HIB),
            card4=kpi_card("⛽", "Flota Combustión (CI)", flota_ci, ACC_CI, BG_CI),
            )

            # 👇 IMPORTANTE: usar el componente HTML para que no se muestre como texto
            st.components.v1.html(kpi_css + cards_html, height=230, scrolling=False)


            


            # ===== Estética (colores "perfectos" previos) + “N°” + anchos =====
            cfg = {
                "header_bg": "#2D5252",   # verde tecsur del header
                "header_fg": "#FFFFFF",
                "family_bg": "#ECFDF5",   # verde muy suave para familia
                "family_fg": "#065F46",
                "cmap": "Blues",          # mapa de calor azul en valores
                "header_font_size": "1.05rem",  # un poquito más grande
            }

            df_show = piv.reset_index().rename(columns={col_fam: "Familia"})
            # Convertimos índice en N°
            df_show = df_show.reset_index(drop=True)
            df_show.insert(0, "N°", df_show.index + 1)

            num_cols = [c for c in df_show.columns if c not in ("N°", "Familia")]

            # Anchos
            W_NUM = 64     # default numéricas
            W_N   = 20     # columna N°
            W_FAM = 220    # columna Familia
            PAD   = "4px 6px"

            sty = (
                df_show.style
                .format("{:,.0f}", subset=num_cols)
                .background_gradient(cmap=cfg["cmap"], subset=num_cols)
                .set_table_styles([
                    # Encabezado (mismos colores de antes + fuente un poco más grande)
                    {"selector": "thead th", "props": [
                        ("background-color", cfg["header_bg"]),
                        ("color", cfg["header_fg"]),
                        ("font-weight", "600"),
                        ("font-size", cfg["header_font_size"]),
                        ("text-align", "center"),
                        ("border", "1px solid #e5e7eb"),
                        ("padding", PAD),
                        ("white-space", "nowrap"),
                    ]},
                    # Celdas
                    {"selector": "tbody td", "props": [
                        ("border", "1px solid #eef2f7"),
                        ("padding", PAD),
                        ("white-space", "nowrap"),
                        ("text-align", "center"),
                    ]},
                    # Primera columna (Familia) con fondo suave y texto verde
                    {"selector": "tbody th", "props": [
                        ("background-color", cfg["family_bg"]),
                        ("color", cfg["family_fg"]),
                        ("border", "1px solid #eef2f7"),
                        ("padding", PAD),
                        ("text-align", "left"),
                        ("font-weight", "600"),
                    ]},
                ])
                .hide(axis="index")  # sin fila extra de índice
            )

            # Atributos de tabla y colgroup para controlar anchos de forma fiable
            tbl_id = "flota_tbl"
            # Fuerza el ancho de la columna "N°" por nombre
            sty = sty.set_properties(
                subset=["N°"],
                **{
                    "width": f"{W_N}px",
                    "min-width": f"{W_N}px",
                    "max-width": f"{W_N}px",
                    "text-align": "center",
                    "padding-left": "4px",
                    "padding-right": "4px",
                    "white-space": "nowrap",
                }
            )




            tbl_id = "flota_tbl"
            sty = sty.set_table_attributes(
                f'id="{tbl_id}" class="{tbl_id}" '
                'style="border-collapse:collapse;width:100%;table-layout:auto;"'
            )


 
      

            tbl_id = "flota_tbl"  # ya lo definiste antes al asignar atributos al styler

   
     
            css = f"""
            <style>
            /* N° = primera columna visible (clase col0 que genera Pandas) */
            #{tbl_id} th.col0, #{tbl_id} td.col0 {{
                width: {W_N}px !important;
                min-width: {W_N}px !important;
                max-width: {W_N}px !important;
                padding-left: 2px !important;
                padding-right: 2px !important;
                text-align: center !important;
                white-space: nowrap !important;
                box-sizing: border-box !important;
            }}

            /* Familia = segunda visible (col1) */
            #{tbl_id} th.col1, #{tbl_id} td.col1 {{
                width: {W_FAM}px !important;
                min-width: {W_FAM}px !important;
                max-width: {W_FAM}px !important;
                text-align: left !important;
                box-sizing: border-box !important;
            }}

            /* Compacta el resto */
            #{tbl_id} th, #{tbl_id} td {{
                padding: {PAD} !important;
                box-sizing: border-box !important;
            }}
            </style>
            """



            html = css + sty.to_html()
            st.components.v1.html(html, height=640, scrolling=True)

        # ============================ SECCIÓN 3 ============================
        st.markdown("### Flota Tecsur — Distribución por Usuario")

        # --- Indicadores (mismo estilo de tarjetas) ---
        kpi_css_usr = """
        <style>
        .kpiu-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px; margin-bottom: 8px; }
        @media (max-width: 1200px) { .kpiu-grid { grid-template-columns: repeat(2, 1fr); } }
        .kpiu {
        position: relative; background: #FFFFFF; border: 1px solid #eef2f7;
        border-radius: 18px; padding: 18px 18px 14px 18px;
        box-shadow: 0 6px 18px rgba(17,24,39,.08);
        }
        .kpiu::before{ content:""; position:absolute; left:0; right:0; top:0; height:4px;
        background: var(--accent, #2D5252); border-top-left-radius:18px; border-top-right-radius:18px; }
        .kpiu-top { display:flex; align-items:center; gap:10px; margin-bottom:6px; }
        .kpiu-icon {
        width:30px; height:30px; border-radius:999px; display:flex; align-items:center; justify-content:center;
        font-size:18px; background: var(--accent-bg, #E6F4F1); color: var(--accent,#2D5252);
        }
        .kpiu-label { font-size: 13px; color: #475569; font-weight: 600; }
        .kpiu-value { font-size: 36px; line-height: 1.0; font-weight: 800; color:#0f172a; letter-spacing: -0.5px; }
        .kpiu-sub { font-size: 12px; color:#64748b; margin-top: 2px; }
        </style>
        """
       
        def kpiu_card(icon, label, value="—", accent="#2D5252", accent_bg="#E6F4F1", sub=""):
            return f"""
            <div class="kpiu notranslate" style="--accent:{accent};--accent-bg:{accent_bg};" translate="no">
            <div class="kpiu-top">
                <div class="kpiu-icon">{icon}</div>
                <div class="kpiu-label"><span class="notranslate" translate="no">{label}</span></div>
            </div>
            <div class="kpiu-value"><span class="notranslate" translate="no">{value}</span></div>
            <div class="kpiu-sub"><span class="notranslate" translate="no">{sub}</span></div>
            </div>
            """


        # --------- TABLA base y métricas ----------
        # Reutilizamos df_flota_parque (hoja "Flota", header=7, desde columna C)
        col_usuario = find_col_in(df_flota_parque, ["usuario", "usuario final", "usuario asignado"])
        col_comb3   = find_col_in(df_flota_parque, ["combustible", "motorizacion", "motorización", "fuel"])
        col_placa3  = find_col_in(df_flota_parque, ["placa"])

        if not (col_usuario and col_comb3):
            st.info("No encontré **Usuario** y/o **Combustible** en la hoja *Flota* para este cuadro.")
        else:
            dff = df_flota_parque.copy()

            # Normalizaciones
            if col_placa3:
                dff["_PLACA_NORM"] = dff[col_placa3].astype(str).str.upper().str.strip()
            else:
                dff["_PLACA_NORM"] = dff.index.astype(str)  # fallback

            dff[col_usuario] = dff[col_usuario].astype(str).str.strip()
            dff[col_comb3]   = (
                dff[col_comb3].astype(str).str.strip().str.upper()
                .str.replace("ELECTRICO", "ELÉCTRICO", regex=False)
                .str.replace("HIBRIDA", "HÍBRIDA", regex=False)
                .str.replace("HÍBRIDO", "HÍBRIDA", regex=False)
                .str.replace("GASOHOL", "GASOLINA", regex=False)
            )

            # Evitar duplicados por placa → nos quedamos con 1 fila por placa (si hay repetidas)
            dff = dff.drop_duplicates(subset=["_PLACA_NORM"])

            # Pivot: filas = Usuario, columnas = Combustible, valores = # placas únicas
            pivu = (
                dff[[col_usuario, col_comb3, "_PLACA_NORM"]]
                .dropna(subset=[col_usuario, col_comb3])
                .pivot_table(
                    index=col_usuario, columns=col_comb3, values="_PLACA_NORM",
                    aggfunc=lambda s: s.nunique(), fill_value=0
                )
            )

            ordered3 = [c for c in ["DIÉSEL", "DIESEL", "ELÉCTRICO", "GASOLINA", "HÍBRIDA"] if c in pivu.columns]
            if "DIESEL" in pivu.columns and "DIÉSEL" not in pivu.columns:
                pivu = pivu.rename(columns={"DIESEL": "DIÉSEL"})
                ordered3 = [c.replace("DIESEL", "DIÉSEL") for c in ordered3]
            other3 = [c for c in pivu.columns if c not in ordered3]
            pivu   = pivu[ordered3 + other3] if ordered3 else pivu

            pivu["Usuario total"] = pivu.sum(axis=1)
            pivu = pivu.sort_values("Usuario total", ascending=False)
            pivu.columns.name = None

            # --- KPIs calculados ---
            total_usuarios = int((pivu.index.astype(str) != "").sum())
            total_veh      = int(pivu["Usuario total"].sum())
            prom_veh_usr   = (total_veh / total_usuarios) if total_usuarios else 0.0

            top_row = pivu["Usuario total"].sort_values(ascending=False).head(1)
            if not top_row.empty:
                top_usuario = str(top_row.index[0])
                top_count   = int(top_row.iloc[0])
                top_label   = f"{top_usuario}"
                top_sub     = f"{top_count} vehículos"
            else:
                top_label, top_sub = "—", ""

            # % EV sobre total (en este cuadro por usuario)
            ev_total = int(pivu["ELÉCTRICO"].sum()) if "ELÉCTRICO" in pivu.columns else 0
            pct_ev   = (ev_total / total_veh * 100.0) if total_veh else 0.0

        
    
            cards_html_usr = f"""
            <div class="kpiu-grid">
            { kpiu_card("👤", "Usuarios", f"{total_usuarios:,}") }
            { kpiu_card("🚗", "Vehículos (total)", f"{total_veh:,}") }
            { kpiu_card("⚡", "% de vehículos eléctricos", f"{pct_ev:,.1f} %", accent="#0EA5A0", accent_bg="#ECFDF5", sub=f"{ev_total} EV de {total_veh}") }
            { kpiu_card("🏆", "Usuario con más vehículos", top_label, sub=top_sub) }
            </div>
            """
            st.components.v1.html(kpi_css_usr + cards_html_usr, height=160, scrolling=False)


                        

            # ===== Estilo de la tabla =====
            cfg3 = {
                "header_bg": "#2D5252",
                "header_fg": "#FFFFFF",
                "cmap": "Blues",
                "header_font_size": "1.05rem",
            }

            df_show3 = pivu.reset_index().rename(columns={col_usuario: "Usuario"})
            df_show3 = df_show3.reset_index(drop=True)
            df_show3.insert(0, "N°", df_show3.index + 1)

            num_cols3 = [c for c in df_show3.columns if c not in ("N°", "Usuario")]
            W_NUM3, W_N3, W_USR3, PAD3 = 64, 20, 260, "4px 6px"

            sty3 = (
                df_show3.style
                .format("{:,.0f}", subset=num_cols3)
                .background_gradient(cmap=cfg3["cmap"], subset=num_cols3)
                .set_table_styles([
                    {"selector": "thead th", "props": [
                        ("background-color", cfg3["header_bg"]),
                        ("color", cfg3["header_fg"]),
                        ("font-weight", "600"),
                        ("font-size", cfg3["header_font_size"]),
                        ("text-align", "center"),
                        ("border", "1px solid #e5e7eb"),
                        ("padding", PAD3),
                        ("white-space", "nowrap"),
                    ]},
                    {"selector": "tbody td", "props": [
                        ("border", "1px solid #eef2f7"),
                        ("padding", PAD3),
                        ("white-space", "nowrap"),
                        ("text-align", "center"),
                    ]},
                ])
                .hide(axis="index")
                .set_properties(subset=["N°"], **{
                    "width": f"{W_N3}px","min-width": f"{W_N3}px","max-width": f"{W_N3}px",
                    "text-align":"center","padding-left":"4px","padding-right":"4px","white-space":"nowrap",
                })
                .set_properties(subset=["Usuario"], **{
                    "width": f"{W_USR3}px","min-width": f"{W_USR3}px","max-width": f"{W_USR3}px",
                    "text-align":"left","font-weight":"600","background-color":"#f0fdf4",
                })
            )

            tbl_id3 = "flota_usuario_tbl"
            sty3 = sty3.set_table_attributes(
                f'id="{tbl_id3}" class="{tbl_id3}" style="border-collapse:collapse;width:100%;table-layout:auto;"'
            )

            css3 = f"""
            <style>
            /* Fuerza N° estrecho (2ª visible al ocultar índice) */
            #{tbl_id3} thead th:nth-child(2),
            #{tbl_id3} tbody td:nth-child(2) {{
                width: {W_N3}px !important; max-width:{W_N3}px !important; min-width:{W_N3}px !important;
                padding-left:6px !important; padding-right:6px !important; text-align:center !important; white-space:nowrap !important;
            }}
            /* Usuario ancho (3ª visible) */
            #{tbl_id3} thead th:nth-child(3),
            #{tbl_id3} tbody td:nth-child(3) {{
                width: {W_USR3}px !important; max-width:{W_USR3}px !important; min-width:{W_USR3}px !important;
                text-align:left !important;
            }}
            #{tbl_id3} th, #{tbl_id3} td {{ padding: {PAD3} !important; box-sizing:border-box; }}
            </style>
            """
         
            st.components.v1.html(css3 + sty3.to_html(), height=420, scrolling=True)

        # ==========================================================

       
        # ============== Flota Tecsur — CR · Marca · Placa (con filtros y KPIs) ==============

        st.markdown('<div style="margin-top:-8px;"></div>', unsafe_allow_html=True)
        st.markdown("### Flota Tecsur — CR · Marca · Placa")

        # --- Carga robusta de la hoja Flota ---
        try:
            flota = read_flota(FLOTA_PATH)  # lee hoja "Flota", header=7 y desde columna C
        except Exception as e:
            flota = pd.DataFrame()
            st.error(f"No pude leer la hoja Flota: {e}")

        if flota.empty:
            st.info("No encontré datos en **Flota** para este cuadro.")
        else:
            # ---------- Helpers ----------
            def pick_col(df, candidates):
                for c in candidates:
                    if c in df.columns:
                        return c
                raise KeyError(f"No se encontró ninguna de las columnas: {candidates}")

            COL_AREA  = pick_col(flota, ["Área Tecsur", "Area Tecsur", "Área", "Area"])
            COL_MOTOR = pick_col(flota, ["Motorización", "Motorizacion"])     # EV / CI
            COL_CR    = pick_col(flota, ["CR Tecsur", "CR", "CR_Tecsur"])
            COL_MARCA = pick_col(flota, ["Marca"])
            COL_PLACA = pick_col(flota, ["Placa Vehículo", "Placa Vehiculo", "Placa"])

            # ---------- Filtros ----------
            areas_opts = ["Todos"] + sorted(flota[COL_AREA].dropna().astype(str).unique())
            moto_opts  = ["Todos"] + sorted(flota[COL_MOTOR].dropna().astype(str).unique())

            col_f1, col_f2, _sp = st.columns([1.2, 1.2, 4.6])
            with col_f1:
                area_sel = st.selectbox("**Área Tecsur**", areas_opts, index=0)
            with col_f2:
                moto_sel = st.selectbox("**Motorización**", moto_opts, index=0)

            f = flota.copy()
            if area_sel != "Todos":
                f = f[f[COL_AREA].astype(str) == str(area_sel)]
            if moto_sel != "Todos":
                f = f[f[COL_MOTOR].astype(str) == str(moto_sel)]

            # ---------- Tabla base ----------
            out = (
                f[[COL_CR, COL_MARCA, COL_PLACA]]
                .dropna(subset=[COL_PLACA])
                .rename(columns={COL_CR: "CR Tecsur", COL_MARCA: "Marca", COL_PLACA: "Placa Vehículo"})
                .copy()
            )
            out["Cuenta de Placa Vehículo"] = 1

            # ---------- KPIs ----------
            total_placas = int(out["Cuenta de Placa Vehículo"].sum()) if not out.empty else 0
            cant_ev = int((f[COL_MOTOR] == "EV").sum())
            cant_ci = int((f[COL_MOTOR] == "CI").sum())
            pct_ev  = (cant_ev / (cant_ev + cant_ci) * 100.0) if (cant_ev + cant_ci) > 0 else 0.0
            marcas_uni = int(out["Marca"].nunique()) if not out.empty else 0
            cr_unicos  = int(out["CR Tecsur"].nunique()) if not out.empty else 0

            # ---------- Indicadores (estilo pantallazo 1) ----------
            kpi_css = """
            <style>
            .kpi-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:22px}
            .kpi-card{background:#fff;border-radius:16px;box-shadow:0 8px 20px rgba(2,25,35,.06);
                        padding:18px 22px;border-top:6px solid var(--acc);}
            .kpi-label{font-size:.95rem;color:#5b6570;display:flex;align-items:center;gap:8px}
            .kpi-value{font-size:40px;line-height:1.1;margin-top:6px;font-weight:700;color:#0f172a}
            .kpi-sub{font-size:.8rem;color:#6b7280;margin-top:2px}
            .kpi-icon{font-size:20px}
            @media (max-width:1200px){.kpi-grid{grid-template-columns:1fr 1fr}.kpi-value{font-size:32px}}
            </style>
            """
            def kpi_card(icon, label, value, acc, sub=None):
                sub_html = f'<div class="kpi-sub">{sub}</div>' if sub else ""
                return f"""
                <div class="kpi-card" style="--acc:{acc}">
                <div class="kpi-label"><span class="kpi-icon">{icon}</span>{label}</div>
                <div class="kpi-value">{value}</div>
                {sub_html}
                </div>
                """

            ACC_TOTAL = "#1c6b63"
            ACC_EV    = "#1c6b63"
            ACC_TAG   = "#1c6b63"
            ACC_CR    = "#1c6b63"

            cards_html = (
                '<div class="kpi-grid">'
                + kpi_card("🚘", "Placas (filtradas)", total_placas, ACC_TOTAL)
                + kpi_card("⚡", "% EV en filtro", f"{pct_ev:.1f} %", ACC_EV, sub=f"{cant_ev} EV de {cant_ev+cant_ci}")
                + kpi_card("🏷️", "Marcas únicas", marcas_uni, ACC_TAG)
                + kpi_card("🏢", "CR únicos", cr_unicos, ACC_CR)
                + '</div>'
            )
            st.components.v1.html(kpi_css + cards_html, height=180, scrolling=False)

            # ---------- Estilo de tabla (pantallazo 2) ----------
            from pandas.io.formats.style import Styler
            def style_grid(df: pd.DataFrame) -> Styler:
                sty = (
                    df.style
                    .set_table_styles([
                        {"selector": "thead th", "props": [
                            ("background-color", "#1F4D4D"),
                            ("color", "#FFFFFF"),
                            ("font-weight", "700"),
                            ("font-size", "1.0rem"),
                            ("text-align", "center"),
                            ("border", "1px solid #dce2e8"),
                            ("padding", "6px 10px"),
                            ("white-space", "nowrap"),
                        ]},
                        {"selector": "tbody td", "props": [
                            ("border", "1px solid #eef2f7"),
                            ("padding", "6px 10px"),
                            ("white-space", "nowrap"),
                            ("text-align", "center"),
                        ]},
                    ])
                    .hide(axis="index")
                )
                if len(df) > 0:
                    sty = sty.apply(
                        lambda s: ["background-color: #f5fbf8" if (s.name % 2 == 0) else ""] * len(s),
                        axis=1
                    )
                return sty

            st.dataframe(style_grid(out), use_container_width=True)

# ---------------------- TAB: Detalle ----------------------
with tab_detalle:
    st.subheader("Flota (hoja 'Flota')")

    try:
        df_flota_det = read_flota(FLOTA_PATH)   # helper que ya tienes (lee hoja "Flota", header=7, desde columna C)
    except Exception as e:
        df_flota_det = pd.DataFrame()
        st.error(f"No pude leer la hoja Flota: {e}")

    if df_flota_det.empty:
        st.info("No encontré datos en **Flota**.")
    else:
        # Detecta columnas clave de forma robusta
        col_placa   = find_col_in(df_flota_det, ["placa vehiculo", "placa vehículo", "placa"])
        col_familia = find_col_in(df_flota_det, ["familia", "segmento", "tipo"])
        col_comb    = find_col_in(df_flota_det, ["combustible", "motorizacion", "motorización", "fuel"])
        col_cliente = find_col_in(df_flota_det, ["cliente asignado", "cliente"])

        # Arma el orden de columnas a mostrar (solo las que existan)
        cols = [c for c in [col_placa, col_familia, col_comb, col_cliente] if c]

        if not cols:
            st.info("No hallé columnas esperadas en **Flota** (Placa, Familia, Combustible, Cliente).")
        else:
            dfv = df_flota_det.copy()

            # Normalizaciones básicas
            if col_placa:
                dfv[col_placa] = dfv[col_placa].map(lambda x: norm_placa(x))
            if col_comb:
                dfv[col_comb] = (
                    dfv[col_comb].astype(str).str.strip().str.upper()
                    .str.replace("ELECTRICO", "ELÉCTRICO", regex=False)
                    .str.replace("HIBRIDA", "HÍBRIDA", regex=False)
                    .str.replace("HÍBRIDO", "HÍBRIDA", regex=False)
                )

            # Vista ordenada y limpia
            df_show = (
                dfv[cols]
                .dropna(subset=[col_placa]) if col_placa else dfv[cols]
            ).drop_duplicates()
            # Renombra amigable
            ren = {}
            if col_placa:   ren[col_placa]   = "Placa"
            if col_familia: ren[col_familia] = "Familia"
            if col_comb:    ren[col_comb]    = "Combustible"
            if col_cliente: ren[col_cliente] = "Cliente Asignado"
            df_show = df_show.rename(columns=ren).sort_values(["Familia","Placa"], na_position="last")

            # ===== Estilos finos: encabezado con color y columnas delgadas =====
            header_bg = "#1f4e78"; header_fg = "#ffffff"
            sty = (
                df_show.style
                .set_table_styles([
                    {"selector": "th", "props": [
                        ("background-color", header_bg),
                        ("color", header_fg),
                        ("text-align", "center"),
                        ("border", "1px solid #e5e7eb"),
                        ("padding", "2px 4px"),       # encabezado más delgado
                    ]},
                    {"selector": "td", "props": [
                        ("border", "1px solid #e5e7eb"),
                        ("padding", "2px 6px"),       # celdas más delgadas
                        ("font-size", "0.85rem"),
                    ]},
                ])
                # anchos por columna (ajústalos a gusto)
                .set_properties(subset=["Placa"], **{"width": "110px", "min-width": "100px", "max-width": "130px", "text-align": "left"})
                .set_properties(subset=["Familia"], **{"width": "210px", "min-width": "180px", "max-width": "260px", "text-align": "left", "font-weight": "600", "background-color": "#f0fdf4"})
                .set_properties(subset=["Combustible"], **{"width": "120px", "min-width": "100px", "max-width": "140px"})
            )

            if "Cliente Asignado" in df_show.columns:
                sty = sty.set_properties(subset=["Cliente Asignado"], **{
                    "width": "180px", "min-width": "160px", "max-width": "240px", "text-align": "left"
                })

            st.dataframe(sty, use_container_width=True, hide_index=True)







# ------------------- TAB: Lista de verificación -------------------
with tab_check:
    st.subheader("Lista de verificación")

    if not INSPECCIONES_PATH.exists():
        st.info("No encontré el archivo **Inspecciones.xlsx**.")
        st.stop()

    # --- Cargas base ---
    df_detalle = read_detalle(INSPECCIONES_PATH)
    df_flota   = read_flota(FLOTA_PATH)
    fechas_map, links_map = read_flota_oper_info(FLOTA_OPER_PATH)

    # --- Detectar columnas clave en Detalle ---
    placa_col = next((c for c in df_detalle.columns if canon(c) in {"placa", "vehiculo", "vehículo", "unidad"}), None)
    fecha_col = next((c for c in df_detalle.columns if canon(c) in {canon("Datos Generales_Fecha"), "fecha"}), None)
    cond_col  = next((c for c in df_detalle.columns if "conductor" in canon(c)), None)
    if not placa_col:
        st.error("No se encontró la columna **Placa** en Inspecciones → Detalle.")
        st.stop()

    # --- Filtrado básico por fecha y placa ---
   
    _det = df_detalle.copy()

    # Fuerza columna de fecha normalizada a date
    if fecha_col:
        _det["_fecha_det"] = pd.to_datetime(_det[fecha_col], errors="coerce").dt.date
    else:
        _det["_fecha_det"] = pd.NaT

    # Diagnóstico rápido (mostrar rango disponible en archivo)
    if fecha_col:
        st.caption(
            f"Rango disponible en Inspecciones.xlsx: "
            f"{_det['_fecha_det'].min()} → {_det['_fecha_det'].max()}"
        )

    # Aplica filtros del panel
    if fecha_col:
        _det = _det[
            (_det["_fecha_det"] >= fecha_ini) &
            (_det["_fecha_det"] <= fecha_fin)
        ]

    if placa_filtro:
        _det = _det[
            _det[placa_col].astype(str).str.contains(placa_filtro.strip(), case=False, na=False)
        ]

  
    # Fallback si quedó vacío: muestra últimos 90 días del archivo
    if _det.empty and fecha_col:
        st.warning("No hay inspecciones en el rango/placa seleccionado. Mostrando últimos 90 días disponibles.")
        _det = df_detalle.copy()
        _det["_fecha_det"] = pd.to_datetime(_det[fecha_col], errors="coerce").dt.date
        cutoff = date.today() - timedelta(days=90)   # ← clave: datetime.timedelta, sin .date()
        _det = _det[_det["_fecha_det"] >= cutoff]

        if _det.empty:
            st.info("No se encontraron inspecciones recientes en el archivo.")
            st.stop()  # ← importante para no seguir con DataFrames vacíos

            

    _det["_placa"] = _det[placa_col].map(norm_placa)


    # --- Rango de columnas por grupo de ítems (coincide con tu hoja Inspecciones) ---
    headers = list(_det.columns)
    def names_in_range(col_a: str, col_b: str) -> list[str]:
        a = excel_col_to_idx(col_a); b = excel_col_to_idx(col_b)
        a = max(0, a); b = min(len(headers)-1, b)
        if a > b: return []
        return [headers[i] for i in range(a, b+1)]

    grupos = {
        "Documentos":  names_in_range("C", "G"),
        "Implementos": names_in_range("H", "R"),
        "Luces":       names_in_range("S", "V"),
        "Neumáticos":  names_in_range("W", "X"),
        "Unidad":      names_in_range("Y", "AR"),
    }

    # --- Mapa Cliente Asignado desde Flota ---
    placa_col_flota = next((c for c in df_flota.columns if canon(c).startswith("placa")), None)
    cliente_col     = next((c for c in df_flota.columns if canon(c) == "cliente asignado"), None)
    map_cliente = {}
    if placa_col_flota and cliente_col:
        dff = df_flota[[placa_col_flota, cliente_col]].dropna(subset=[placa_col_flota]).copy()
        dff["_placa"] = dff[placa_col_flota].map(norm_placa)
        map_cliente = dff.drop_duplicates("_placa").set_index("_placa")[cliente_col].to_dict()

    # --- Última inspección por placa ---
    ult = _det.sort_values(fecha_col if fecha_col else _det.index).drop_duplicates("_placa", keep="last")

    # ============ Construcción de datasets (resumen + fallas) ============
    resumen_rows = []
    fallas_rows  = []

    for _, row in ult.iterrows():
        p = row["_placa"]
        registro = {"Placa": p, "Cliente asignado": map_cliente.get(p, "")}

        for cat, cols_cat in grupos.items():
            vals = [str(row.get(c, "")).strip().lower() for c in cols_cat]
            hay_mal = any(v == "mal" for v in vals)
            hay_ok  = any(v in ("ok", "conforme") for v in vals)

            if hay_mal:
                registro[cat] = "❌"
            elif hay_ok:
                registro[cat] = "✅"
            else:
                registro[cat] = ""

            for c in cols_cat:
                v = str(row.get(c, "")).strip().lower()
                if v == "mal":
                    fallas_rows.append({
                        "Placa": p,
                        "Categoría": cat,
                        "Ítem": c,
                        "Fecha checklist": fechas_map.get(p, ""),
                        "Conductor": (row.get(cond_col, "") if cond_col else ""),
                        "Link": links_map.get(p) or _guess_pdf_link(p)

                    })

        resumen_rows.append(registro)

    df_resumen = pd.DataFrame(resumen_rows).sort_values("Placa").reset_index(drop=True)
    df_fallas  = pd.DataFrame(fallas_rows)

    # =================== KPIs arriba ===================
        # =================== KPIs + Semáforo ===================
    cat_cols = ["Documentos","Implementos","Luces","Neumáticos","Unidad"]

    # Semáforo global por placa
    df_resumen["Estado general"] = df_resumen.apply(_estado_general, cat_cols=cat_cols, axis=1)

    total_uni = df_resumen.shape[0]
    con_falla = df_resumen.set_index("Placa")[cat_cols].eq("❌").any(axis=1).sum()
    pct_falla = (con_falla / total_uni * 100) if total_uni else 0.0
    ult_prom  = pd.to_datetime([x for x in fechas_map.values() if x], format="%d/%m/%Y", errors="coerce")

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Inspeccionadas", f"{total_uni:,}")
    k2.metric("Unidades con fallas", f"{con_falla:,}")
    k3.metric("% con al menos 1 falla", f"{pct_falla:,.1f}%")
    k4.metric("Fecha promedio última inspección",
            "" if ult_prom.isna().all() else ult_prom.dropna().mean().date().strftime("%d/%m/%Y"))

    st.markdown("---")

    # =================== Filtros rápidos ===================
    cat_opts = ["Todas"] + cat_cols
    c1, c2 = st.columns([2,2])
    with c1:
        cat_sel = st.selectbox("Filtrar por categoría", cat_opts, index=0)
    with c2:
        solo_fallas = st.checkbox("Mostrar solo vehículos con fallas", value=False)

    # =================== Resumen gráfico (% fallas por categoría) ===================
    if total_uni > 0:
        _base = (df_resumen[cat_cols]
                .replace({"❌": 1, "✅": 0, "": 0})
                .apply(pd.to_numeric, errors="coerce")
                .fillna(0))
        pct = (_base.sum(axis=0).astype(float) / float(total_uni) * 100.0).reset_index()
        pct.columns = ["Categoría","% con fallas"]
        # Paleta por categoría (azul, naranja, verde, gris, morado)
        palette_domain = ["Implementos", "Unidad", "Luces", "Neumáticos", "Documentos"]
        palette_range  = ["#1f77b4", "#ff7f0e", "#2ca02c", "#7f7f7f", "#9467bd"]

        chart = (
            alt.Chart(pct)
            .mark_bar()
            .encode(
                x=alt.X("% con fallas:Q", title="% con fallas"),
                y=alt.Y("Categoría:N", sort="-x", title="Categoría"),
                color=alt.Color(
                    "Categoría:N",
                    scale=alt.Scale(domain=palette_domain, range=palette_range),
                    legend=None,
                ),
                tooltip=[
                    alt.Tooltip("Categoría:N", title="Categoría"),
                    alt.Tooltip("% con fallas:Q", title="% con fallas", format=".1f"),
                ],
            )
            .properties(height=220)
        )

        st.altair_chart(chart, use_container_width=True)

    st.markdown("---")

    # =================== Tabla principal (resumen + semáforo) ===================
    _tb = df_resumen.copy()
    if solo_fallas:
        _mask = _tb[cat_cols].eq("❌").any(axis=1)
        _tb = _tb[_mask]
    if cat_sel != "Todas":
        _tb = _tb[_tb[cat_sel] == "❌"]

    cols_show = ["Placa","Cliente asignado","Estado general"] + cat_cols
    st.markdown("#### Resumen por placa")
    st.dataframe(_tb[cols_show], use_container_width=True, hide_index=True)

        # =================== Vista expandible por placa ===================
    st.markdown("#### Detalle por placa")

    # helper local: cuenta ítems MAL/OK en una fila de checklist
    def _items_mal_ok(row, cols):
        vals = [str(row.get(c, "")).strip().lower() for c in cols]
        mal = [c for c, v in zip(cols, vals) if v == "mal"]
        ok  = [c for c, v in zip(cols, vals) if v in ("ok", "conforme")]
        return mal, ok

    placas_list = _tb["Placa"].tolist() if not _tb.empty else []
    if placas_list:
        for p in placas_list:
            # última inspección disponible para la placa dentro del rango seleccionado
            ult_p = (
                _det[_det["_placa"] == p]
                .sort_values(fecha_col if fecha_col else _det.index)
                .tail(1)
            )
            if ult_p.empty:
                continue

            last = ult_p.iloc[0]
            subt = f"Conductor: {last.get(cond_col, '') or '—'}"
           
            raw = links_map.get(p)
            # si está vacío o vino el texto del hipervínculo ("Ver PDF"), usa fallback
            if not raw or raw.strip().lower() == "ver pdf":
                raw = _guess_pdf_link(p)

            link_directo = sp_web_view(sp_normalize(raw))
            link_pdf     = f"[Ver inspección]({link_directo})"
            link_carpeta = f"[Abrir carpeta]({sp_web_view(sp_normalize(BASE_INSP_FOLDER_VIEW))})"




            with st.expander(f"➕ {p} · {subt} · {link_pdf} · {link_carpeta}", expanded=False):

                # Resumen por categoría (OK/MAL/Total) de la última inspección
                rows_cat = []
                for cat, cols_cat in grupos.items():
                    if not cols_cat:
                        continue
                    mal, ok = _items_mal_ok(last, cols_cat)
                    rows_cat.append({
                        "Categoría": cat,
                        "OK": len(ok),
                        "MAL": len(mal),
                        "Total ítems": len(cols_cat),
                    })
                df_cat_res = pd.DataFrame(rows_cat)

                c1, c2 = st.columns([1, 2], vertical_alignment="top")

                with c1:
                    st.markdown("**Resumen por categoría**")
                    if not df_cat_res.empty:
                        df_cat_res = df_cat_res.sort_values(["MAL", "Categoría"], ascending=[False, True])
                        st.dataframe(df_cat_res, use_container_width=True, hide_index=True)
                    else:
                        st.caption("Sin datos de categorías para esta placa.")

                with c2:
                    # Fallas (ítems 'MAL') y link al PDF
                    df_p = df_fallas[df_fallas["Placa"] == p].copy()
                    if cat_sel != "Todas":
                        df_p = df_p[df_p["Categoría"] == cat_sel]
                    if df_p.empty:
                        st.caption("Sin fallas registradas para esta placa con los filtros aplicados.")
                    else:
                        st.markdown("**Fallas de la última inspección**")
                        cols_mostrar = ["Categoría", "Ítem", "Fecha checklist", "Conductor"]  # sin "Link"
                        df_p = df_p[cols_mostrar].sort_values(["Categoría", "Ítem"])
                        st.dataframe(df_p, use_container_width=True, hide_index=True)

                        # --- Visor / descarga del PDF del checklist ---
                        st.markdown("---")
                        st.markdown("**Checklist (PDF)**")

                        # PDF local (opcional) y URL SharePoint
                        try:
                            pdf_local = find_local_pdf(p)   # si tienes este helper; si no, quedará en None
                        except NameError:
                            pdf_local = None

                        
                    
                        # URL del PDF desde el Excel (o fallback si vino "Ver PDF"/vacío)
                        raw = links_map.get(p)
                        if not raw or str(raw).strip().lower() == "ver pdf":
                            raw = _guess_pdf_link(p)

                        # URL final lista para abrir en el navegador
                        link_directo = sp_web_view(sp_normalize(raw))


                        b1, b2 = st.columns([1, 2])

                                                
                        with b1:
                            # Botón que abre *fuera* de la app
                            st.markdown(
                                f'<a href="{link_directo}" target="_blank" rel="noopener" '
                                'style="background:#0a64a4;color:white;padding:8px 16px;border-radius:8px;'
                                'text-decoration:none;font-weight:600;display:inline-block;width:100%;text-align:center;">'
                                '🌐 Ver Check list en SharePoint</a>',
                                unsafe_allow_html=True
                            )


                            # (deja igual tu bloque de descarga local)
                            #if pdf_local is not None:
                             #   try:
                              #      pdf_bytes = pdf_local.read_bytes()
                               #     st.download_button(
                                #        "⬇️ Descargar PDF (local)",
                                 #       data=pdf_bytes,
                                  #      file_name=pdf_local.name,
                                   #     mime="application/pdf",
                                    #    use_container_width=True,
                                    #)
                                #except Exception as e:
                                 #   st.caption(f"No pude leer el archivo local: {e}")
                            #else:
                             #   st.caption("No encontré el PDF local de esta placa.")



                       
                        #with b2:
                         #   # visor embebido desactivado
                          #  st.caption("Abre el PDF desde el botón de la izquierda.")

                           

        else:
            st.caption("No hay detalle de fallas para mostrar con los filtros aplicados.")

        
        # --- Visor / botón para abrir el PDF en SharePoint ---
              
        #pdf_url_raw = links_map.get(p)
        #if not pdf_url_raw or str(pdf_url_raw).strip().lower() == "ver pdf":
         #   pdf_url_raw = _guess_pdf_link(p)

        #pdf_url_abs = sp_normalize(pdf_url_raw)
        #pdf_url_web = sp_web_view(pdf_url_abs) or sp_web_view(sp_normalize(BASE_INSP_FOLDER_VIEW))

        #st.markdown(
         #   f'<a href="{pdf_url_web}" target="_blank" rel="noopener" '
          #  'style="display:inline-flex;align-items:center;justify-content:center;'
           # 'width:100%;padding:.5rem 1rem;border-radius:.5rem;border:1px solid #d0d7de;text-decoration:none;">'
            #'🌐 Abrir en SharePoint</a>',
            #unsafe_allow_html=True
        #)




    #st.link_button("🌐 Abrir en SharePoint", pdf_url_web, use_container_width=True)


    # =================== Ítems de falla más frecuentes (Detalle de fallas frecuentes) ===================

    # Pareto de fallas por ítem
    # ============================
    st.markdown("#### Pareto de fallas por ítem")
    if not df_fallas.empty:
        # Trae selecciones desde la sidebar (o usa defaults)
        cat_sel     = st.session_state.get("cat_sel", "Todas")
        familia_sel = st.session_state.get("familia_sel", "Todas")
        cliente_sel = st.session_state.get("cliente_sel", "Todos")

        # Si arriba ya filtraste una tabla _tb con placas visibles, úsala como universo
        placas_sel = None
        try:
            placas_sel = set(_tb["Placa"].dropna().astype(str).str.strip())
        except Exception:
            pass

        _f = df_fallas.copy()

        # Normaliza columnas base a string por si acaso
        for c in ["Placa", "Categoría", "Familia", "Cliente Asignado", "Ítem"]:
            if c in _f.columns:
                _f[c] = _f[c].astype(str).str.strip()

        # Filtros defensivos (solo si la columna existe)
        if cat_sel != "Todas" and "Categoría" in _f.columns:
            _f = _f[_f["Categoría"] == cat_sel]

        if familia_sel != "Todas" and "Familia" in _f.columns:
            _f = _f[_f["Familia"] == familia_sel]

        if cliente_sel != "Todos" and "Cliente Asignado" in _f.columns:
            _f = _f[_f["Cliente Asignado"] == cliente_sel]

        if placas_sel is not None and "Placa" in _f.columns and len(placas_sel) > 0:
            _f = _f[_f["Placa"].isin(placas_sel)]

        # Agrupa y ordena
        if not _f.empty and {"Categoría", "Ítem"}.issubset(_f.columns):
            freq = (
                _f.groupby(["Categoría", "Ítem"]).size()
                .reset_index(name="Nº de fallas")
                .sort_values("Nº de fallas", ascending=False)
            )

            if not freq.empty:
                st.dataframe(freq, use_container_width=True,hide_index=True)
            else:
                st.info("No hay fallas que cumplan los filtros actuales.")
        else:
            st.info("No se encontraron columnas esperadas ('Categoría', 'Ítem') en df_fallas.")
    else:
        st.info("No hay datos de fallas para mostrar.")



# ---------------------- TAB: Exportar ----------------------
with tab_export:
    st.subheader("Descargas")

    # Cargamos la misma fuente que usa la pestaña Detalle
    try:
        df_flota_exp = read_flota(FLOTA_PATH)   # hoja "Flota", header=7, desde columna C
    except Exception as e:
        df_flota_exp = pd.DataFrame()
        st.error(f"No pude leer la hoja Flota: {e}")

    if df_flota_exp.empty:
        st.info("No encontré datos en **Flota** para exportar.")
    else:
        # Columnas clave (mismo criterio que en Detalle) + nuevas: Marca, Modelo, Usuario Final
        col_placa     = find_col_in(df_flota_exp, ["placa vehiculo", "placa vehículo", "placa"])
        col_familia   = find_col_in(df_flota_exp, ["familia", "segmento", "tipo"])
        col_comb      = find_col_in(df_flota_exp, ["combustible", "motorizacion", "motorización", "fuel"])
        col_cliente   = find_col_in(df_flota_exp, ["cliente asignado", "cliente"])
        col_marca     = find_col_in(df_flota_exp, ["marca"])
        col_modelo    = find_col_in(df_flota_exp, ["modelo", "versión", "version"])
        col_usuariofn = find_col_in(df_flota_exp, ["usuario final", "usuario", "usuario asignado", "asignado a"])

        # Armamos el set de columnas a exportar
        cols = [c for c in [
            col_placa, col_familia, col_comb,
            col_marca, col_modelo, col_usuariofn,
            col_cliente
        ] if c]

        if not cols or not col_placa:
            st.info("No hallé columnas esperadas en **Flota** (Placa, Familia, Combustible, Marca, Modelo, Usuario Final) para exportar.")
        else:
            dfv = df_flota_exp.copy()

            # Normalizaciones idénticas a Detalle
            if col_placa:
                dfv[col_placa] = dfv[col_placa].map(lambda x: norm_placa(x))
            if col_comb:
                dfv[col_comb] = (
                    dfv[col_comb].astype(str).str.strip().str.upper()
                    .str.replace("ELECTRICO", "ELÉCTRICO", regex=False)
                    .str.replace("HIBRIDA", "HÍBRIDA", regex=False)
                    .str.replace("HÍBRIDO", "HÍBRIDA", regex=False)
                )

            # Selección y limpieza
            df_show = (
                dfv[cols]
                .dropna(subset=[col_placa])     # Placa obligatoria para exportar
                .drop_duplicates()
            )

            # Renombrados bonitos
            ren = {}
            if col_placa:     ren[col_placa]     = "Placa"
            if col_familia:   ren[col_familia]   = "Familia"
            if col_comb:      ren[col_comb]      = "Combustible"
            if col_marca:     ren[col_marca]     = "Marca"
            if col_modelo:    ren[col_modelo]    = "Modelo"
            if col_usuariofn: ren[col_usuariofn] = "Usuario Final"
            if col_cliente:   ren[col_cliente]   = "Cliente Asignado"
            df_show = df_show.rename(columns=ren)

            # Orden sugerido de columnas
            preferred_order = [
                "Placa", "Familia", "Combustible",
                "Marca", "Modelo", "Usuario Final",
                "Cliente Asignado",
            ]
            final_cols = [c for c in preferred_order if c in df_show.columns] + \
                         [c for c in df_show.columns if c not in preferred_order]
            df_show = df_show[final_cols].sort_values(["Familia", "Placa"], na_position="last")

            # ===== Botones de descarga =====
            from io import BytesIO
            from datetime import date
            hoy = date.today().strftime("%Y_%m_%d")
            fname_csv   = f"Flota_Tecsur_{hoy}.csv"
            fname_excel = f"Flota_Tecsur_{hoy}.xlsx"

            # CSV (utf-8-sig para Excel)
            csv_bytes = df_show.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "⬇️ Exportar CSV (Flota)",
                data=csv_bytes,
                file_name=fname_csv,
                mime="text/csv",
                use_container_width=False,
            )

            # Excel
            bio = BytesIO()
            with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
                df_show.to_excel(writer, index=False, sheet_name="Flota")
                wb  = writer.book
                ws  = writer.sheets["Flota"]

                # Anchos compactos por defecto
                for i, col in enumerate(df_show.columns):
                    # Ajuste base compacto (puedes bajar a 14–16 si quieres aún más angosto)
                    ws.set_column(i, i, 18)

                # Encabezado con color
                header_fmt = wb.add_format({
                    "bold": True, "font_color": "white",
                    "align": "center", "valign": "vcenter",
                    "bg_color": "#1f4e78", "border": 1
                })
                for col_idx, col_name in enumerate(df_show.columns):
                    ws.write(0, col_idx, col_name, header_fmt)

            bio.seek(0)
            st.download_button(
                "⬇️ Exportar Excel (Flota)",
                data=bio.getvalue(),
                file_name=fname_excel,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=False,
            )
